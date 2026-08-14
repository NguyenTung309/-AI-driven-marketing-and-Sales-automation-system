"""Kiem tra format sheet Unit Test (5.1) so voi sheet mau methodName1.

Sheet 5.1 la ma tran ngang nen khong dung duoc compare_format.py (von so theo
dong). O day anh xa CA dong lan cot cua sheet ket qua ve sheet mau:

- dong: 1..7 va 4 dong Result anh xa 1:1; trong tung band, dong dau va dong cuoi
  anh xa 1:1, cac dong giua anh xa ve dong giua cuoi cung cua band mau
- cot: cac cot truoc cot cuoi anh xa 1:1, cot cuoi anh xa ve cot cuoi cua mau,
  cac cot chen them anh xa ve cot giap bien cua mau

Khi kich thuoc trung mau, anh xa tro thanh 1:1 -> phep so tro thanh so TUNG O,
day la phep kiem chat nhat (dung cho ban replica).

    python scripts/testdoc/compare_unit.py --base "docs/test/Report5.1_Unit Test.xlsx" \
        --target out.xlsx --pair methodName1=IsAllowedBaseUrl
"""

from __future__ import annotations

import argparse
import dataclasses
import sys

import openpyxl
from openpyxl.utils import get_column_letter

from compare_format import (cell_style_key, column_key, raw_row_dims, size_equal,
                            style_diff)

WIDTH_TOL = 0.02


@dataclasses.dataclass(frozen=True)
class Layout:
    last_col: int
    bands: tuple[tuple[int, int], ...]
    result_row: int
    max_row: int
    max_col: int


def read_layout(ws) -> Layout:
    merged = next((m for m in ws.merged_cells.ranges
                   if m.min_row == 3 and m.min_col == 3), None)
    if merged is None:
        raise ValueError(f"{ws.title}: khong thay merge 'Test requirement' o C3")
    starts, result_row = [], None
    for row in range(8, 500):
        label = str(ws.cell(row, 1).value or "").strip().lower()
        if label == "result":
            result_row = row
            break
        if label:
            starts.append(row)
    if result_row is None or not starts:
        raise ValueError(f"{ws.title}: thieu dong 'Result' hoac band dieu kien")
    bands = tuple((start, (starts[i + 1] if i + 1 < len(starts) else result_row) - 1)
                  for i, start in enumerate(starts))
    return Layout(last_col=merged.max_col, bands=bands, result_row=result_row,
                  max_row=result_row + 4, max_col=merged.max_col)


def row_map(base: Layout, target: Layout) -> dict[int, int]:
    mapping = {row: row for row in range(1, 8)}
    for index, (start, end) in enumerate(target.bands):
        if index >= len(base.bands):
            raise ValueError(f"ket qua co {len(target.bands)} band, mau chi co {len(base.bands)}")
        base_start, base_end = base.bands[index]
        # Dong giu lai tu mau phai giong CHINH NO (anh xa 1:1); chi dong chen
        # them moi doi chieu voi dong mau (dong giua cuoi cung cua band).
        for offset in range(end - start):
            mapping[start + offset] = min(base_start + offset, base_end - 1)
        mapping[end] = base_end  # dong cuoi mang vien duoi cua band
    for offset in range(5):
        mapping[target.result_row + offset] = base.result_row + offset
    return mapping


def col_map(base: Layout, target: Layout) -> dict[int, int]:
    if target.last_col == base.last_col:
        return {col: col for col in range(1, target.max_col + 1)}
    mapping = {col: col for col in range(1, base.last_col)}
    for col in range(base.last_col, target.last_col):
        mapping[col] = base.last_col - 1  # cot chen them -> cot giap bien cua mau
    mapping[target.last_col] = base.last_col
    return mapping


def dv_map(ws) -> dict[tuple[int, int], str]:
    rules: dict[tuple[int, int], str] = {}
    for rule in ws.data_validations.dataValidation:
        for chunk in rule.sqref.ranges:
            for row in range(chunk.min_row, chunk.max_row + 1):
                for col in range(chunk.min_col, chunk.max_col + 1):
                    rules[(row, col)] = str(rule.formula1)
    return rules


def merge_set(ws, rmap: dict, cmap: dict, max_row: int) -> set:
    mapped = set()
    for rng in ws.merged_cells.ranges:
        if rng.min_row > max_row:
            continue
        try:
            mapped.add((rmap[rng.min_row], cmap[rng.min_col],
                        rmap[rng.max_row], cmap[rng.max_col]))
        except KeyError:
            mapped.add(("ngoai vung", str(rng)))
    return mapped


def compare_cells(base_ws, target_ws, rmap, cmap, target: Layout, diffs: list) -> None:
    base_dv, target_dv = dv_map(base_ws), dv_map(target_ws)
    for row in range(1, target.max_row + 1):
        for col in range(1, target.max_col + 1):
            base_row, base_col = rmap[row], cmap[col]
            addr = f"{get_column_letter(col)}{row}"
            proto = f"{get_column_letter(base_col)}{base_row}"
            problems = style_diff(cell_style_key(base_ws.cell(base_row, base_col)),
                                  cell_style_key(target_ws.cell(row, col)),
                                  allow_date_numfmt=True)
            if problems:
                diffs.append(f"o {addr} (mau {proto}): {', '.join(problems)}")
            if base_dv.get((base_row, base_col)) != target_dv.get((row, col)):
                diffs.append(f"o {addr} (mau {proto}): data validation "
                             f"{base_dv.get((base_row, base_col))!r} -> "
                             f"{target_dv.get((row, col))!r}")


def compare_sheet(base_ws, target_ws, base_dims, target_dims) -> list[str]:
    base, target = read_layout(base_ws), read_layout(target_ws)
    rmap, cmap = row_map(base, target), col_map(base, target)
    diffs: list[str] = []

    compare_cells(base_ws, target_ws, rmap, cmap, target, diffs)

    base_rows, target_rows = base_dims.get(base_ws.title, {}), target_dims.get(target_ws.title, {})
    for row in range(1, target.max_row + 1):
        want, got = base_rows.get(rmap[row], (None, False)), target_rows.get(row, (None, False))
        if not size_equal(want, got, 0.5):
            diffs.append(f"dong {row} (mau {rmap[row]}): chieu cao/an {want} -> {got}")

    base_cols, target_cols = column_key(base_ws, base.max_col), column_key(target_ws, target.max_col)
    for col in range(1, target.max_col + 1):
        letter, proto = get_column_letter(col), get_column_letter(cmap[col])
        want, got = base_cols[proto], target_cols[letter]
        if not size_equal(want, got, WIDTH_TOL):
            diffs.append(f"cot {letter} (mau {proto}): rong/an {want} -> {got}")

    base_merges = merge_set(base_ws, {r: r for r in range(1, base.max_row + 1)},
                            {c: c for c in range(1, base.max_col + 1)}, base.max_row)
    # Mau co mot merge le D17:E17 nam giua band Condition. Khi band ngan hon mau,
    # dong 17 khong con ton tai nen khong the doi hoi merge do - bo qua cac merge
    # cua mau nam tren dong ma ket qua khong con anh xa toi.
    reachable = set(rmap.values())
    base_merges = {item for item in base_merges if item[0] in reachable}
    target_merges = merge_set(target_ws, rmap, cmap, target.max_row)
    for missing in sorted(base_merges - target_merges, key=str):
        diffs.append(f"thieu merge {missing}")
    for extra in sorted(target_merges - base_merges, key=str):
        diffs.append(f"thua merge {extra}")

    if base_ws.freeze_panes != target_ws.freeze_panes:
        diffs.append(f"freeze panes {base_ws.freeze_panes} -> {target_ws.freeze_panes}")
    return diffs


def main() -> int:
    ap = argparse.ArgumentParser(description="So format sheet Unit Test voi sheet mau")
    ap.add_argument("--base", required=True)
    ap.add_argument("--target", required=True)
    ap.add_argument("--pair", action="append", required=True,
                    help="sheet_mau=sheet_ket_qua (lap lai duoc)")
    ap.add_argument("--limit", type=int, default=40)
    args = ap.parse_args()

    base_wb = openpyxl.load_workbook(args.base)
    target_wb = openpyxl.load_workbook(args.target)
    base_dims, target_dims = raw_row_dims(args.base), raw_row_dims(args.target)

    total = 0
    for pair in args.pair:
        base_name, target_name = pair.split("=", 1)
        diffs = compare_sheet(base_wb[base_name], target_wb[target_name], base_dims, target_dims)
        total += len(diffs)
        print(f"== {base_name} -> {target_name}: {len(diffs)} khac biet")
        for line in diffs[:args.limit]:
            print(f"   {line}")
        if len(diffs) > args.limit:
            print(f"   ... con {len(diffs) - args.limit} dong nua")

    print("\nFORMAT KHOP 100% - khong co khac biet" if total == 0
          else f"\nCON {total} KHAC BIET FORMAT")
    return 0 if total == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
