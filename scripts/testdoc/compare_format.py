"""So sanh FORMAT giua file mau va file ket qua (bo qua noi dung).

Muc dich: chung minh output ra dung format cua mau 100%.
Chi doc bang openpyxl, khong bao gio save.

So sanh nhung gi:
  1. thuoc tinh sheet: freeze panes, print area, print titles, page setup
  2. do rong / an cua tung cot
  3. chieu cao / an cua tung dong trong vung header
  4. style tung o vung header (font, fill, border, alignment, number format)
  5. style dong du lieu: moi dong du lieu cua target phai khop style dong mau
  6. merge trong vung header: khop tuyet doi
  7. data validation: bo rule (type + formula1) va do phu tren cac cot ket qua
  8. (tuy chon) quet #REF! trong target

Cach dung:
  python compare_format.py --base "docs/test/Report5.3_System Test.xlsx" \
      --target out.xlsx --same-name \
      --pair "Workflow Name1=WF01 Onboarding" --header-rows 11 --data-row 12 \
      --dv-cols F,I,L --check-ref
"""

from __future__ import annotations

import argparse
import json
import sys
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def raw_row_dims(path: str) -> dict[str, dict[int, tuple]]:
    """{ten sheet: {dong: (chieu cao dat tay, an)}} doc thang tu XML.

    Bat buoc phai doc XML: openpyxl tra ve customHeight = 'co ht hay khong',
    khong phai thuoc tinh that. Dong co ht nhung customHeight=0 la chieu cao
    Excel tu tinh (cache), khong phai format do nguoi dat -> khong tinh la khac.
    """
    out: dict[str, dict[int, tuple]] = {}
    with zipfile.ZipFile(path) as zf:
        book = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        target = {rel.get("Id"): rel.get("Target") for rel in rels}
        for sheet in book.find(f"{_NS}sheets"):
            part = target[sheet.get(f"{_REL_NS}id")]
            part = part.lstrip("/") if part.startswith("/") else f"xl/{part}"
            dims: dict[int, tuple] = {}
            data = ET.fromstring(zf.read(part)).find(f"{_NS}sheetData")
            for row in data if data is not None else []:
                custom = row.get("customHeight") in ("1", "true")
                height = row.get("ht")
                dims[int(row.get("r"))] = (
                    round(float(height), 2) if custom and height else None,
                    row.get("hidden") in ("1", "true"),
                )
            out[sheet.get("name")] = dims
    return out


# --------------------------------------------------------------------------- #
# Chuan hoa style
# --------------------------------------------------------------------------- #
def color_key(color):
    if color is None:
        return None
    kind = getattr(color, "type", None)
    value = getattr(color, kind, None) if kind else None
    if not isinstance(value, (str, int, float, type(None))):
        value = str(value)
    tint = getattr(color, "tint", 0.0) or 0.0
    return (kind, value, round(float(tint), 4))


def font_key(font):
    if font is None:
        return None
    return (
        font.name,
        float(font.sz) if font.sz is not None else None,
        bool(font.b),
        bool(font.i),
        font.u,
        bool(font.strike),
        color_key(font.color),
        font.vertAlign,
    )


def fill_key(fill):
    if fill is None:
        return None
    return (
        fill.fill_type,
        color_key(getattr(fill, "fgColor", None)),
        color_key(getattr(fill, "bgColor", None)),
    )


def border_key(border):
    """Canh vien. Side vang mat, Side rong va Side co mau nhung khong co style
    deu la 'khong ke vien' -> quy ve cung mot gia tri, neu khong Excel luu lai
    file la sinh ra hang nghin khac biet ao.
    """
    out = []
    for side_name in ("left", "right", "top", "bottom", "diagonal"):
        side = getattr(border, side_name, None) if border is not None else None
        style = getattr(side, "style", None) if side is not None else None
        out.append((style, color_key(side.color)) if style else (None, None))
    return tuple(out)


def align_key(alignment):
    if alignment is None:
        return None
    return (
        alignment.horizontal,
        alignment.vertical,
        bool(alignment.wrap_text),
        alignment.text_rotation,
        alignment.indent,
        bool(alignment.shrink_to_fit),
    )


def cell_style_key(cell):
    return {
        "font": font_key(cell.font),
        "fill": fill_key(cell.fill),
        "border": border_key(cell.border),
        "align": align_key(cell.alignment),
        "numfmt": cell.number_format,
    }


def is_date_numfmt(numfmt):
    if not numfmt or numfmt == "General":
        return False
    lowered = numfmt.lower()
    return any(token in lowered for token in ("yy", "mm/", "/mm", "dd", "mmm"))


def style_diff(base_style, target_style, allow_date_numfmt=False):
    changed = []
    for key in base_style:
        if base_style[key] == target_style[key]:
            continue
        # Hai truong hop ve number format ngay thang KHONG tinh la lech:
        #  - mau de General, ket qua co dinh dang ngay (o ngay bat buoc phai co)
        #  - ca hai deu la dinh dang ngay: chinh Excel doi 'dd/MM/yyyy' thanh
        #    dinh dang ngay dung san theo locale moi lan mo ra luu lai
        if (
            key == "numfmt"
            and allow_date_numfmt
            and is_date_numfmt(target_style[key])
            and (base_style[key] in (None, "General") or is_date_numfmt(base_style[key]))
        ):
            continue
        changed.append(key)
    return changed


def size_equal(a, b, tol):
    """So sanh do rong cot / chieu cao dong voi dung sai (Excel lam tron theo font)."""
    (a_size, a_hidden), (b_size, b_hidden) = a, b
    if bool(a_hidden) != bool(b_hidden):
        return False
    if a_size is None or b_size is None:
        return a_size == b_size
    return abs(float(a_size) - float(b_size)) <= tol


# --------------------------------------------------------------------------- #
# Trich thuoc tinh sheet
# --------------------------------------------------------------------------- #
def page_setup_key(ws):
    ps = ws.page_setup
    fit = getattr(ws.sheet_properties.pageSetUpPr, "fitToPage", None)
    return {
        "orientation": ps.orientation,
        "paperSize": ps.paperSize,
        "scale": ps.scale,
        "fitToWidth": ps.fitToWidth,
        "fitToHeight": ps.fitToHeight,
        "fitToPage": fit,
        "print_area": ws.print_area,
        "print_title_rows": ws.print_title_rows,
        "print_title_cols": ws.print_title_cols,
        "freeze_panes": ws.freeze_panes,
        "showGridLines": ws.sheet_view.showGridLines,
    }


def column_key(ws, max_col):
    out = {}
    for idx in range(1, max_col + 1):
        letter = get_column_letter(idx)
        dim = ws.column_dimensions.get(letter)
        if dim is None:
            out[letter] = (None, False)
        else:
            width = round(float(dim.width), 2) if dim.width is not None else None
            out[letter] = (width, bool(dim.hidden))
    return out


def row_key(raw_dims, rows):
    return {r: raw_dims.get(r, (None, False)) for r in rows}


def merges_in_rows(ws, max_row):
    out = set()
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, mx_row = range_boundaries(str(rng))
        if mx_row <= max_row:
            out.add(str(rng))
    return out


def dv_rules(ws):
    rules = set()
    for dv in ws.data_validations.dataValidation:
        rules.add((dv.type, dv.operator, str(dv.formula1), str(dv.formula2)))
    return rules


def dv_coverage(ws):
    covered = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type is None:
            continue
        for rng in str(dv.sqref).split():
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            if max_row - min_row > 5000:  # tranh no bo nho voi vung ca cot
                max_row = min_row + 5000
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    covered.add((r, c))
    return covered


def data_rows_of(ws, data_row, id_col=1, limit=2000):
    rows = []
    for r in range(data_row, min(ws.max_row, data_row + limit) + 1):
        if ws.cell(r, id_col).value not in (None, ""):
            rows.append(r)
    return rows


def is_group_row(ws, row, span_cols=(2, 3, 4, 5)):
    """Dong nhom scenario: chi co gia tri o cot A, cac cot con lai rong."""
    if ws.cell(row, 1).value in (None, ""):
        return False
    return all(ws.cell(row, c).value in (None, "") for c in span_cols)


def blank_formatted_rows(ws, data_row, limit=40):
    """Dong cua mau da ke san vien nhung chua co chu.

    Mau 5.4/5.5 dung HAI bien the cho vung du lieu: dong vi du (in nghieng,
    chu xam - y la "thay toi di") va dong trong da ke vien san cho du lieu
    that. Neu chi lay style tu dong CO CHU thi bo nghiem thu se bat moi dong
    du lieu that phai in nghieng xam nhu dong vi du, tuc bat sai y mau.
    """
    rows = []
    for r in range(data_row, min(ws.max_row, data_row + limit) + 1):
        cell = ws.cell(r, 1)
        if cell.value not in (None, ""):
            continue
        border = cell.border
        if any(getattr(getattr(border, side, None), "style", None)
               for side in ("left", "right", "top", "bottom")):
            rows.append(r)
    return rows


def style_prototypes(ws, data_row, max_col, limit=40):
    """Bo style hop le cho tung cot, lay tu vung du lieu cua sheet mau.

    Mau co the dung nhieu bien the style cho cac dong vi du khac nhau, nen
    chap nhan bat ky bien the nao xuat hien trong mau - ke ca bien the nam o
    cac dong trong da ke vien san (xem blank_formatted_rows).
    """
    protos = {c: [] for c in range(1, max_col + 1)}
    for r in data_rows_of(ws, data_row, limit=limit):
        if is_group_row(ws, r):
            continue
        for c in range(1, max_col + 1):
            protos[c].append(cell_style_key(ws.cell(r, c)))
    # Dong mau goc luon la mot bien the hop le, ke ca voi sheet khong co dong
    # du lieu nao khac de lay mau (vi du Cover).
    for c in protos:
        protos[c].append(cell_style_key(ws.cell(data_row, c)))
    # THEM (khong thay the) bien the "dong trong da ke vien san" cua mau.
    for r in blank_formatted_rows(ws, data_row, limit=limit):
        for c in range(1, max_col + 1):
            protos[c].append(cell_style_key(ws.cell(r, c)))
    return protos


# --------------------------------------------------------------------------- #
# So sanh 1 cap sheet
# --------------------------------------------------------------------------- #
def compare_sheet(base_ws, target_ws, header_rows, data_row, dv_cols, max_data_rows,
                  size_tol=0.2, base_dims=None, target_dims=None):
    base_dims = base_dims if base_dims is not None else {}
    target_dims = target_dims if target_dims is not None else {}
    diffs = []
    tag = f"{base_ws.title} -> {target_ws.title}"

    # 1. thuoc tinh sheet
    b_page, t_page = page_setup_key(base_ws), page_setup_key(target_ws)
    for key in b_page:
        if key == "print_area":
            continue  # print area doi theo so dong du lieu, kiem rieng ben duoi
        if b_page[key] != t_page[key]:
            diffs.append(f"[{tag}] page/{key}: mau={b_page[key]!r} ket qua={t_page[key]!r}")
    if bool(b_page["print_area"]) != bool(t_page["print_area"]):
        diffs.append(f"[{tag}] print_area: mau={b_page['print_area']!r} ket qua={t_page['print_area']!r}")

    max_col = max(base_ws.max_column, 18)  # toi thieu toi cot R (nguon data validation)

    # 2. cot
    b_cols, t_cols = column_key(base_ws, max_col), column_key(target_ws, max_col)
    for letter in b_cols:
        if not size_equal(b_cols[letter], t_cols[letter], size_tol):
            diffs.append(f"[{tag}] cot {letter}: mau={b_cols[letter]} ket qua={t_cols[letter]}")

    # 3. dong vung header
    header_range = range(1, header_rows + 1)
    b_rows, t_rows = row_key(base_dims, header_range), row_key(target_dims, header_range)
    for r in header_range:
        if not size_equal(b_rows[r], t_rows[r], size_tol):
            diffs.append(f"[{tag}] dong {r}: mau={b_rows[r]} ket qua={t_rows[r]}")

    # 4. style o vung header
    for r in header_range:
        for c in range(1, max_col + 1):
            b_style = cell_style_key(base_ws.cell(r, c))
            t_style = cell_style_key(target_ws.cell(r, c))
            changed = style_diff(b_style, t_style, allow_date_numfmt=True)
            if changed:
                addr = f"{get_column_letter(c)}{r}"
                for key in changed:
                    diffs.append(
                        f"[{tag}] style {addr}/{key}: mau={b_style[key]!r} ket qua={t_style[key]!r}"
                    )

    # 5. style dong du lieu: moi dong du lieu target phai khop MOT trong cac
    #    bien the style ma sheet mau dang dung
    protos = style_prototypes(base_ws, data_row, max_col)
    group_protos = [
        cell_style_key(base_ws.cell(r, 1))
        for r in data_rows_of(base_ws, data_row, limit=40)
        if is_group_row(base_ws, r)
    ]
    checked = 0
    for r in data_rows_of(target_ws, data_row):
        checked += 1
        if checked > max_data_rows:
            break
        if is_group_row(target_ws, r):
            t_style = cell_style_key(target_ws.cell(r, 1))
            if group_protos and t_style not in group_protos:
                changed = style_diff(group_protos[0], t_style)
                for key in changed:
                    diffs.append(
                        f"[{tag}] style dong nhom A{r}/{key}: "
                        f"mau={group_protos[0][key]!r} ket qua={t_style[key]!r}"
                    )
            continue
        for c in range(1, max_col + 1):
            t_style = cell_style_key(target_ws.cell(r, c))
            if t_style in protos[c]:
                continue
            if any(not style_diff(p, t_style, allow_date_numfmt=True) for p in protos[c]):
                continue
            changed = style_diff(protos[c][0], t_style, allow_date_numfmt=True)
            addr = f"{get_column_letter(c)}{r}"
            for key in changed:
                diffs.append(
                    f"[{tag}] style dong du lieu {addr}/{key}: "
                    f"mau={protos[c][0][key]!r} ket qua={t_style[key]!r}"
                )

    # 6. merge vung header
    b_merge = merges_in_rows(base_ws, header_rows - 1)
    t_merge = merges_in_rows(target_ws, header_rows - 1)
    for missing in sorted(b_merge - t_merge):
        diffs.append(f"[{tag}] thieu merge {missing}")
    for extra in sorted(t_merge - b_merge):
        diffs.append(f"[{tag}] merge thua {extra}")

    # 7. data validation
    b_dv, t_dv = dv_rules(base_ws), dv_rules(target_ws)
    for missing in sorted(b_dv - t_dv, key=str):
        diffs.append(f"[{tag}] thieu rule data validation {missing}")
    for extra in sorted(t_dv - b_dv, key=str):
        diffs.append(f"[{tag}] rule data validation la {extra}")
    if dv_cols:
        # chi doi hoi dropdown o cot ma CHINH MAU co dropdown: cung mot bo cot
        # duoc dung cho moi sheet, trong khi Cover khong phai bang test case
        base_covered = dv_coverage(base_ws)
        base_cols = {
            col_idx
            for col_idx in dv_cols
            if any(
                (r, col_idx) in base_covered
                for r in data_rows_of(base_ws, data_row)
                if not is_group_row(base_ws, r)
            )
        }
        covered = dv_coverage(target_ws)
        for r in data_rows_of(target_ws, data_row):
            if is_group_row(target_ws, r):
                continue
            for col_idx in sorted(base_cols):
                if (r, col_idx) not in covered:
                    addr = f"{get_column_letter(col_idx)}{r}"
                    diffs.append(f"[{tag}] o {addr} khong co dropdown")

    return diffs


def scan_ref_errors(wb_values):
    hits = []
    for ws in wb_values.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and "#REF!" in value:
                    hits.append(f"[{ws.title}] {cell.coordinate}: {value}")
    return hits


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="So sanh format giua file mau va file ket qua")
    ap.add_argument("--base", required=True, help="file mau")
    ap.add_argument("--target", required=True, help="file ket qua")
    ap.add_argument("--pair", action="append", default=[], help='"sheet mau=sheet ket qua", lap nhieu lan')
    ap.add_argument("--same-name", action="store_true", help="tu ghep cac sheet trung ten")
    ap.add_argument("--header-rows", type=int, default=11, help="so dong vung header (mac dinh 11)")
    ap.add_argument("--data-row", type=int, default=12, help="dong du lieu mau (mac dinh 12)")
    ap.add_argument("--dv-cols", default="", help='cot phai co dropdown, vi du "F,I,L"')
    ap.add_argument("--max-data-rows", type=int, default=400, help="so dong du lieu kiem toi da moi sheet")
    ap.add_argument("--size-tol", type=float, default=0.2, help="dung sai do rong cot / chieu cao dong")
    ap.add_argument("--check-ref", action="store_true", help="quet #REF! trong file ket qua")
    ap.add_argument("--json", dest="as_json", action="store_true")
    args = ap.parse_args()

    base_wb = openpyxl.load_workbook(args.base)
    target_wb = openpyxl.load_workbook(args.target)

    pairs = []
    for item in args.pair:
        if "=" not in item:
            print(f"--pair sai dinh dang: {item}", file=sys.stderr)
            return 2
        b, t = item.split("=", 1)
        pairs.append((b.strip(), t.strip()))
    if args.same_name:
        for name in base_wb.sheetnames:
            if name in target_wb.sheetnames and (name, name) not in pairs:
                pairs.append((name, name))

    dv_cols = []
    for token in filter(None, (x.strip() for x in args.dv_cols.split(","))):
        dv_cols.append(openpyxl.utils.column_index_from_string(token))

    base_dims, target_dims = raw_row_dims(args.base), raw_row_dims(args.target)

    diffs = []
    for base_name, target_name in pairs:
        if base_name not in base_wb.sheetnames:
            diffs.append(f"[mau] khong co sheet {base_name!r}")
            continue
        if target_name not in target_wb.sheetnames:
            diffs.append(f"[ket qua] khong co sheet {target_name!r}")
            continue
        diffs += compare_sheet(
            base_wb[base_name],
            target_wb[target_name],
            args.header_rows,
            args.data_row,
            dv_cols,
            args.max_data_rows,
            args.size_tol,
            base_dims.get(base_name, {}),
            target_dims.get(target_name, {}),
        )

    if args.check_ref:
        diffs += [f"#REF! {hit}" for hit in scan_ref_errors(target_wb)]

    if args.as_json:
        print(json.dumps({"pairs": pairs, "diffs": diffs}, ensure_ascii=False, indent=2))
    else:
        print(f"Cap sheet so sanh: {len(pairs)}")
        for b, t in pairs:
            print(f"  {b} -> {t}")
        if diffs:
            print(f"\nKHAC BIET FORMAT: {len(diffs)}")
            for line in diffs[:200]:
                print("  " + line)
            if len(diffs) > 200:
                print(f"  ... con {len(diffs) - 200} dong nua")
        else:
            print("\nFORMAT KHOP 100% - khong co khac biet")

    return 1 if diffs else 0


if __name__ == "__main__":
    sys.exit(main())
